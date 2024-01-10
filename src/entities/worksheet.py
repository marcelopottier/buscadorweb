import openpyxl
from src.entities.proccess import Proccess

class Workbook:
    def __init__(self):
        self.workbook = openpyxl.load_workbook('DEFINIR PLANILHA')
    def createWorkbookPage(workbook, Proccess):
        if Proccess.pri in workbook.sheetnames:
            workbookPage = workbook[proccessId]
        else:
            workbookPage = workbook.create_sheet(proccessId)

    def defineWorkbook(workbookPage, Proccess):
        workbookPage['A1'].value = "Número do Processo"
        workbookPage['A2'].value = 
        workbookPage['B1'].value = "Classe do Processo"
        workbookPage['B2'].value = proccess_class
        workbookPage['C1'].value = "Assunto do Processo"
        workbookPage['C2'].value = proccess_subject
        workbookPage['D1'].value = "Partes do Processo"
        max_row = len(lista_partes_process) + 1
        for index, row in enumerate(workbookPage.iter_rows(min_row=2, max_row=max_row, min_col=4, max_col=4)):
            for cell in row:
                cell.value = lista_partes_process[index]
        workbookPage['E1'].value = "Foro do Processo"
        workbookPage['E2'].value = proccess_foro
        workbookPage['F1'].value = "Vara do Processo"
        workbookPage['F2'].value = proccess_vara
        workbookPage['G1'].value = "Juiz do Processo"
        workbookPage['G2'].value = proccess_judge
        workbookPage['H1'].value = "Distribuição do Processo"
        workbookPage['H2'].value = proccess_distruicao
        workbookPage['I1'].value = "Controle do Processo"
        workbookPage['I2'].value = proccess_controle
        workbookPage['J1'].value = "Área do Processo"
        workbookPage['J2'].value = proccess_area
        workbookPage['K1'].value = "Valor do Processo"
        workbookPage['K2'].value = proccess_valor
        workbookPage['L1'].value = "Data Movimentação"
        workbookPage['M1'].value = "Descrição Movimentação"

        for index, row in enumerate(workbookPage.iter_rows(min_row=2, max_row=len(lista_data_movimentacoes), min_col=12, max_col=12)):
            for cell in row:
                cell.value = lista_data_movimentacoes[index]

        for index, row in enumerate(workbookPage.iter_rows(min_row=2, max_row=len(lista_desc_movimentacoes), min_col=13, max_col=13)):
            for cell in row:
                cell.value = lista_desc_movimentacoes[index]

        workbook.save('dados.xlsx')